"""Telco bill processing workflows for Singtel and M1."""

from __future__ import annotations

import os
import re
import shutil
import calendar
from datetime import datetime
from typing import Dict, List, Tuple

import pandas as pd
import openpyxl


def get_current_month_year() -> Tuple[str, str, str]:
    """Get current month name, 2-digit year, and 4-digit year."""
    now = datetime.now()
    month_name = now.strftime('%b')  # Sep, Oct, etc.
    year_short = now.strftime('%y')  # 25
    year_full = now.strftime('%Y')  # 2025
    return month_name, year_short, year_full


def process_singtel_bills(pdf1_path: str, pdf2_path: str, igs32_path: str, cnt35_path: str) -> Dict[str, str]:
    """
    Process Singtel bills by renaming and copying to specified paths.
    
    Args:
        pdf1_path: Path to first PDF (will be renamed to IGS SIP)
        pdf2_path: Path to second PDF (will be renamed to IGS Telco)
        igs32_path: Base path for Singtel-IGS.32
        cnt35_path: Base path for Singtel-CNT.35
    
    Returns:
        Dictionary containing paths to saved files
    """
    month_name, year_short, _ = get_current_month_year()
    
    # Generate new filenames
    sip_filename = f"IGS SIP {month_name}{year_short}Bill.pdf"
    telco_filename = f"IGS Telco {month_name}{year_short}Bill.pdf"
    
    # Ensure directories exist
    os.makedirs(igs32_path, exist_ok=True)
    os.makedirs(cnt35_path, exist_ok=True)
    
    # Copy files to both paths with new names
    sip_igs32 = os.path.join(igs32_path, sip_filename)
    sip_cnt35 = os.path.join(cnt35_path, sip_filename)
    telco_igs32 = os.path.join(igs32_path, telco_filename)
    telco_cnt35 = os.path.join(cnt35_path, telco_filename)
    
    shutil.copy2(pdf1_path, sip_igs32)
    shutil.copy2(pdf1_path, sip_cnt35)
    shutil.copy2(pdf2_path, telco_igs32)
    shutil.copy2(pdf2_path, telco_cnt35)
    
    return {
        'sip_igs32': sip_igs32,
        'telco_igs32': telco_igs32,
        'sip_cnt35': sip_cnt35,
        'telco_cnt35': telco_cnt35,
        'igs32_path': igs32_path
    }


def process_m1_bill(pdf_path: str, igs32_path: str, cnt35_path: str) -> Dict[str, str]:
    """
    Process M1 bill by renaming and copying to specified paths.
    
    Args:
        pdf_path: Path to M1 PDF
        igs32_path: Base path for M1-IGS.32
        cnt35_path: Base path for M1-CNT.35
    
    Returns:
        Dictionary containing paths to saved files
    """
    month_name, year_short, _ = get_current_month_year()
    
    # Generate new filename
    m1_filename = f"IGS-{month_name.upper()}{year_short}-M1-Bill.pdf"
    
    # Ensure directories exist
    os.makedirs(igs32_path, exist_ok=True)
    os.makedirs(cnt35_path, exist_ok=True)
    
    # Copy files to both paths
    m1_igs32 = os.path.join(igs32_path, m1_filename)
    m1_cnt35 = os.path.join(cnt35_path, m1_filename)
    
    shutil.copy2(pdf_path, m1_igs32)
    shutil.copy2(pdf_path, m1_cnt35)
    
    return {
        'm1_igs32': m1_igs32,
        'm1_cnt35': m1_cnt35
    }


def find_total_row(ws) -> int:
    """Find the row number containing 'Total' in column B."""
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=2).value
        if cell_value and str(cell_value).strip().lower() == 'total':
            return row
    raise ValueError("Could not find 'Total' row in Excel file")


def calculate_price_without_gst(price_with_gst: float) -> float:
    """Calculate price without GST (deduct 9%)."""
    # Price without GST = Price with GST / 1.09
    return round(price_with_gst / 1.09, 2)


def get_previous_month_amount(ws, total_row: int) -> Tuple[str, float]:
    """Get the previous month (row before Total) amount for reference."""
    if total_row <= 2:
        return "N/A", 0.0
    
    prev_row = total_row - 1
    month_col_value = ws.cell(row=prev_row, column=2).value  # Column B (M1 SIM Only Plan)
    price_col_value = ws.cell(row=prev_row, column=3).value  # Column C (Price)
    
    month_str = str(month_col_value) if month_col_value else "N/A"
    price_val = float(price_col_value) if price_col_value else 0.0
    
    return month_str, price_val


def update_m1_excel(excel_path: str, user_amount: float) -> Tuple[str, float]:
    """
    Update M1 Excel file with new monthly amount.
    
    Args:
        excel_path: Path to IGS-Summarize-M1-Bill Excel file
        user_amount: Monthly amount (with GST) provided by user
    
    Returns:
        Tuple of (previous_month_info, previous_amount) for display
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Find Total row
    total_row = find_total_row(ws)
    
    # Get previous month info for reference
    prev_month_str, prev_amount = get_previous_month_amount(ws, total_row)
    
    # Calculate price without GST
    price_without_gst = calculate_price_without_gst(user_amount)
    
    # Get current month and year
    month_name, year_short, _ = get_current_month_year()
    
    # Insert new row above Total
    ws.insert_rows(total_row)
    
    # Get the row number for the new entry (should be same as old total_row after insertion)
    new_row = total_row
    
    # Calculate No. value (previous row's No. + 1)
    prev_no = ws.cell(row=new_row - 1, column=1).value
    new_no = int(prev_no) + 1 if prev_no else 1
    
    # Fill in data: No, Month-Year, Price (with GST), Price (without GST)
    ws.cell(row=new_row, column=1, value=new_no)  # Column A: No
    ws.cell(row=new_row, column=2, value=f"{month_name}-{year_short}")  # Column B: M1 SIM Only Plan
    ws.cell(row=new_row, column=3, value=user_amount)  # Column C: Price
    ws.cell(row=new_row, column=4, value=price_without_gst)  # Column D: Price without GST
    
    # Update Total row formulas to include the new row
    # After insertion, Total row is now at total_row + 1
    new_total_row = total_row + 1
    
    # Update formulas in columns C and D (Price columns)
    for col in [3, 4]:  # Column C (Price) and D (Price without GST)
        cell = ws.cell(row=new_total_row, column=col)
        if cell.value and isinstance(cell.value, str) and '=' in str(cell.value):
            # Cell contains a formula
            formula = str(cell.value)
            # Update SUM formula range to include the new row
            # Convert formula like =SUM(C2:C27) to =SUM(C2:C28)
            # Match patterns like C2:C27 or D2:D27
            pattern = r'([A-Z]+)(\d+):([A-Z]+)(\d+)'
            
            def update_range(match):
                col_start = match.group(1)
                row_start = match.group(2)
                col_end = match.group(3)
                row_end = int(match.group(4))
                # Extend the range to include the new row (which is new_total_row - 1)
                new_row_end = new_total_row - 1
                return f"{col_start}{row_start}:{col_end}{new_row_end}"
            
            updated_formula = re.sub(pattern, update_range, formula)
            cell.value = updated_formula
    
    # Save workbook
    wb.save(excel_path)
    wb.close()
    
    return prev_month_str, prev_amount


def update_both_m1_excels(igs32_excel: str, cnt35_excel: str, user_amount: float) -> Dict[str, any]:
    """
    Update both M1 Excel files (IGS.32 and CNT.35 paths).
    
    Returns:
        Dictionary with update results
    """
    results = {}
    
    # Update IGS.32 Excel
    try:
        prev_month_32, prev_amount_32 = update_m1_excel(igs32_excel, user_amount)
        results['igs32'] = {
            'success': True,
            'prev_month': prev_month_32,
            'prev_amount': prev_amount_32
        }
    except Exception as e:
        results['igs32'] = {
            'success': False,
            'error': str(e)
        }
    
    # Update CNT.35 Excel
    try:
        prev_month_35, prev_amount_35 = update_m1_excel(cnt35_excel, user_amount)
        results['cnt35'] = {
            'success': True,
            'prev_month': prev_month_35,
            'prev_amount': prev_amount_35
        }
    except Exception as e:
        results['cnt35'] = {
            'success': False,
            'error': str(e)
        }
    
    return results

