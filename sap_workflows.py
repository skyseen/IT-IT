"""SAP workflow helpers: data extraction, preview dialog, and email triggers."""

from __future__ import annotations

import os
import re
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import pandas as pd
from PySide6 import QtCore, QtGui, QtWidgets

from config_manager import get_path, set_path
from email_service import send_sap_creation_email
from openpyxl import load_workbook
from openpyxl.styles import Alignment


SAP_COLUMNS: List[str] = [
    "帳號類型（Account Type）",
    "帳號名稱（Account Name）",
    "費用代碼（Expense Code）",
    "Name",
    "聯繫電話（Contact Phone）",
    "部門（Department）",
    "工號（Employee No）",
    "郵箱（E-mail）",
    "帳號Role（Account Role）",
    "其他說明（Other Description）",
    "CM remark",
    "SR V9 file",
    "STATUS",
]

SAP_COLUMN_ALIASES: Dict[str, List[str]] = {
    "帳號類型（Account Type）": [
        "帳號類型(accounttype)",
        "accounttype",
        "type",
    ],
    "帳號名稱（Account Name）": [
        "帳號名稱(accountname)",
        "帳號名稱(account name)",
        "accountname",
        "帳號名稱",
    ],
    "費用代碼（Expense Code）": [
        "費用代碼(expensecode)",
        "費用代碼(expense code)",
        "expensecode",
        "費用代碼",
    ],
    "Name": [
        "name",
        "username",
        "fullname",
    ],
    "聯繫電話（Contact Phone）": [
        "聯繫電話(contactphone)",
        "聯繫電話(contact phone)",
        "contactphone",
        "電話",
        "phone",
    ],
    "部門（Department）": [
        "部門(department)",
        "department",
        "部門",
    ],
    "工號（Employee No）": [
        "工號(employeeno)",
        "工號(employee no)",
        "employeeno",
        "employeeid",
        "employee no",
        "employeeid",
        "empid",
    ],
    "郵箱（E-mail）": [
        "郵箱(e-mail)",
        "郵箱(email)",
        "email",
        "郵箱",
    ],
    "帳號Role（Account Role）": [
        "帳號role(accountrole)",
        "帳號role(account role)",
        "accountrole",
        "role",
    ],
    "其他說明（Other Description）": [
        "其他說明(otherdescription)",
        "其他說明(other description)",
        "otherdescription",
        "remarks",
        "備註",
    ],
    "CM remark": [
        "cmremark",
        "cm備註",
        "cm remark",
    ],
    "SR V9 file": [
        "srv9file",
        "sr v9 file",
        "srv9",
    ],
}


def _normalize_column_name(name: Optional[str]) -> str:
    if not name:
        return ""
    normalized = str(name)
    normalized = normalized.replace(" ", "").replace("\u3000", "")
    normalized = normalized.replace("（", "(").replace("）", ")")
    normalized = normalized.replace("【", "(").replace("】", ")")
    normalized = normalized.lower()
    return normalized


def _canonical_column(name: Optional[str]) -> Optional[str]:
    normalized = _normalize_column_name(name)
    if not normalized:
        return None
    for canonical, aliases in SAP_COLUMN_ALIASES.items():
        canonical_norm = _normalize_column_name(canonical)
        if normalized == canonical_norm or normalized in aliases:
            return canonical
    # If not matched, return original stripped string
    return str(name).strip() if name else None


def _build_normalized_row(row: pd.Series) -> Dict[str, any]:
    norm_map: Dict[str, any] = {}
    for col in row.index:
        norm_key = _normalize_column_name(col)
        if not norm_key:
            continue
        norm_map[norm_key] = row[col]
    return norm_map


def _get_value_from_row(norm_row: Dict[str, any], canonical_key: str) -> str:
    aliases = SAP_COLUMN_ALIASES.get(canonical_key, [])
    candidates = [_normalize_column_name(canonical_key)] + aliases
    for alias in candidates:
        if alias in norm_row:
            value = norm_row[alias]
            if pd.isna(value):
                continue
            return str(value).strip()
    return ""


def _get_sheet_headers(ws) -> List[str]:
    headers: List[str] = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value).strip() if cell.value is not None else "")
    return headers


def _get_sheet_canonical_headers(ws) -> List[str]:
    headers = _get_sheet_headers(ws)
    canonical_headers: List[str] = []
    for header in headers:
        canonical = _canonical_column(header) or header
        canonical_headers.append(canonical)
    return canonical_headers


def _get_employee_ids_from_sheet(ws) -> List[str]:
    """Extract all employee IDs from a worksheet (works with read-only mode)."""
    headers = _get_sheet_headers(ws)
    canonical_headers = _get_sheet_canonical_headers(ws)
    emp_index: Optional[int] = None
    
    # Find the Employee No column index
    for idx, canonical in enumerate(canonical_headers):
        if _normalize_column_name(canonical) == _normalize_column_name("工號（Employee No）"):
            emp_index = idx
            break
    
    if emp_index is None:
        return []
    
    # Use iter_rows to iterate through rows (compatible with read-only mode)
    ids: List[str] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        # Skip header row (min_row=2), get value from employee column
        if row_idx >= 0 and emp_index < len(row):
            value = str(row[emp_index]).strip() if row[emp_index] is not None else ""
            if value and value.lower() != "none":
                ids.append(value.upper())
    
    return ids


def _build_row_values(canonical_headers: List[str], row_data: Dict[str, str]) -> List[str]:
    values: List[str] = []
    for header in canonical_headers:
        canonical = _canonical_column(header) or header
        key = canonical
        if canonical in SAP_COLUMNS:
            key = canonical
        values.append(row_data.get(key, ""))
    return values


@dataclass
class ParsedSapData:
    rows_to_append: List[Dict[str, str]]
    already_created: List[str]
    other_desc_map: Dict[str, str]


def load_consolidated_dataframe(cons_path: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Load both sheets from consolidated Excel file.
    
    Returns:
        Tuple of (sr_v9_df, ly_v10_df)
    """
    if os.path.exists(cons_path):
        try:
            # Try to load both sheets
            sr_v9_df = pd.read_excel(cons_path, sheet_name="SR V9 file", engine="openpyxl")
        except:
            sr_v9_df = pd.DataFrame(columns=SAP_COLUMNS)
        
        try:
            ly_v10_df = pd.read_excel(cons_path, sheet_name="LY V10 file", engine="openpyxl")
        except:
            ly_v10_df = pd.DataFrame(columns=SAP_COLUMNS)
        
        return sr_v9_df, ly_v10_df
    
    return pd.DataFrame(columns=SAP_COLUMNS), pd.DataFrame(columns=SAP_COLUMNS)


def get_all_existing_employees(cons_path: str) -> List[str]:
    """
    Get all existing employee IDs from all sheets using openpyxl directly.
    Checks SR V9 file and/or LY V10 file sheets (handles if only one exists).
    
    Returns:
        List of all employee IDs found in the consolidated Excel
    """
    if not os.path.exists(cons_path):
        return []
    
    wb = load_workbook(cons_path, read_only=True)
    existing_emp = []
    
    # Check both possible sheet names (may have one or both)
    for sheet_name in ["SR V9 file", "LY V10 file"]:
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        emp_ids = _get_employee_ids_from_sheet(ws)
        existing_emp.extend(emp_ids)
    
    wb.close()
    return existing_emp


def append_to_consolidated(cons_path: str, rows: List[Dict[str, str]]) -> None:
    """
    Append new rows to the appropriate sheet.
    Preserves original Excel formatting, colors, and column widths.
    Applies center alignment to new cells.
    
    Logic:
    - If "SR V9 file" sheet exists: append rows with SR V9 value there
    - If "LY V10 file" sheet exists: append rows without SR V9 value there
    - If only one sheet exists: append all rows to that sheet
    """
    # Load existing workbook to preserve formatting
    wb = load_workbook(cons_path)
    
    # Helper to append rows with alignment
    def append_rows_to_sheet(ws, rows_data):
        # Get sheet headers to determine column order
        canonical_headers = _get_sheet_canonical_headers(ws)
        
        for row_data in rows_data:
            # Build row values in the exact order of sheet columns
            values = _build_row_values(canonical_headers, row_data)
            
            # Append the row
            start_row = ws.max_row + 1
            ws.append(values)
            
            # Apply center alignment to all cells in the new row
            for col_idx in range(1, len(values) + 1):
                cell = ws.cell(row=start_row, column=col_idx)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Check which sheets exist
    has_sr_v9 = "SR V9 file" in wb.sheetnames
    has_ly_v10 = "LY V10 file" in wb.sheetnames
    
    if has_sr_v9 and has_ly_v10:
        # Both sheets exist - separate rows based on SR V9 file column
        sr_v9_rows = []
        ly_v10_rows = []
        
        for row in rows:
            sr_v9_value = str(row.get("SR V9 file", "")).strip()
            if sr_v9_value and sr_v9_value.lower() != "nan":
                sr_v9_rows.append(row)
            else:
                ly_v10_rows.append(row)
        
        if sr_v9_rows:
            append_rows_to_sheet(wb["SR V9 file"], sr_v9_rows)
        if ly_v10_rows:
            append_rows_to_sheet(wb["LY V10 file"], ly_v10_rows)
    
    elif has_ly_v10:
        # Only LY V10 sheet exists - append all rows there
        append_rows_to_sheet(wb["LY V10 file"], rows)
    
    elif has_sr_v9:
        # Only SR V9 sheet exists - append all rows there
        append_rows_to_sheet(wb["SR V9 file"], rows)
    
    else:
        raise ValueError("No valid sheet found in consolidated Excel (expected 'SR V9 file' or 'LY V10 file')")
    
    # Save workbook - this preserves all formatting, colors, column widths
    wb.save(cons_path)


def parse_user_excel(user_df: pd.DataFrame, existing_emp: List[str]) -> ParsedSapData:
    existing_set = set(str(emp).strip().upper() for emp in existing_emp if str(emp).strip())

    rows_to_append: List[Dict[str, str]] = []
    already_created: List[str] = []
    other_desc_map: Dict[str, str] = {}

    for _, row in user_df.iterrows():
        # Build normalized row for flexible column matching
        norm_row = _build_normalized_row(row)
        
        # Get employee number using normalized matching
        emp_no = _get_value_from_row(norm_row, "工號（Employee No）")
        if not emp_no:
            continue
        
        # Check if employee already exists (case-insensitive)
        if emp_no.upper() in existing_set:
            already_created.append(emp_no)
            continue

        # Build new row using normalized column matching for all fields
        new_row = {
            "帳號類型（Account Type）": _get_value_from_row(norm_row, "帳號類型（Account Type）"),
            "帳號名稱（Account Name）": _get_value_from_row(norm_row, "帳號名稱（Account Name）"),
            "費用代碼（Expense Code）": _get_value_from_row(norm_row, "費用代碼（Expense Code）"),
            "Name": _get_value_from_row(norm_row, "Name"),
            "聯繫電話（Contact Phone）": _get_value_from_row(norm_row, "聯繫電話（Contact Phone）"),
            "部門（Department）": _get_value_from_row(norm_row, "部門（Department）"),
            "工號（Employee No）": emp_no,
            "郵箱（E-mail）": _get_value_from_row(norm_row, "郵箱（E-mail）"),
            "帳號Role（Account Role）": _get_value_from_row(norm_row, "帳號Role（Account Role）"),
            "其他說明（Other Description）": _get_value_from_row(norm_row, "其他說明（Other Description）"),
            "CM remark": _get_value_from_row(norm_row, "CM remark"),
            "SR V9 file": _get_value_from_row(norm_row, "SR V9 file"),
        }
        rows_to_append.append(new_row)
        other_desc_map[emp_no] = str(new_row["其他說明（Other Description）"]).strip()

    return ParsedSapData(rows_to_append, already_created, other_desc_map)


def _extract_follow_text(other_desc_map: Dict[str, str], employee_nos: List[str]) -> str:
    if not employee_nos:
        return ""
    pattern = re.compile(r"(SGP\d+)", re.IGNORECASE)
    for emp in employee_nos:
        match = pattern.search(other_desc_map.get(emp, ""))
        if match:
            return f"權限跟隨{match.group(1)}即可"
    return ""


def prompt_ticket_details(parent: QtWidgets.QWidget) -> Tuple[str, str]:
    """Prompt user for ticket image and number using PySide6 dialogs."""
    # File dialog for ticket image
    default_ticket_dir = get_path("sap_ticket_image_dir")
    if not default_ticket_dir or not os.path.isdir(default_ticket_dir):
        default_ticket_dir = ""
    
    ticket_img_path, _ = QtWidgets.QFileDialog.getOpenFileName(
        parent,
        "Select Ticket Image File",
        default_ticket_dir,
        "Image files (*.png *.jpg *.jpeg *.bmp *.gif);;All files (*.*)"
    )
    
    if ticket_img_path:
        set_path("sap_ticket_image_dir", os.path.dirname(ticket_img_path))
    
    # Input dialog for ticket number
    ticket_no, ok = QtWidgets.QInputDialog.getText(
        parent,
        "Ticket Number",
        "Enter ticket number (e.g. S0000YE9G):"
    )
    
    if not ok or not ticket_no:
        raise ValueError("Ticket number is required")
    
    return ticket_no, ticket_img_path


class SapPreviewDialog(QtWidgets.QDialog):
    """SAP Account Creation Preview Dialog using PySide6."""
    
    def __init__(
        self,
        rows_to_append: List[Dict[str, str]],
        already_created: List[str],
        cons_path: str,
        user_excel_path: str,
        other_desc_map: Dict[str, str],
        parent: Optional[QtWidgets.QWidget] = None
    ):
        super().__init__(parent)
        self.rows_to_append = rows_to_append
        self.already_created = already_created
        self.cons_path = cons_path
        self.user_excel_path = user_excel_path
        self.other_desc_map = other_desc_map
        
        # Check for empty descriptions
        self.empty_other_desc = []
        for row in rows_to_append:
            emp_no = row.get("工號（Employee No）", "")
            other_desc = str(row.get("其他說明（Other Description）", "")).strip()
            if not other_desc or other_desc.lower() == "nan":
                self.empty_other_desc.append(emp_no)
        
        self._setup_ui()
    
    def _setup_ui(self):
        self.setWindowTitle("SAP_PREVIEW_MODULE")
        self.setMinimumSize(1200, 800)
        
        # Dark theme styling
        self.setStyleSheet("""
            QDialog {
                background-color: #0d1117;
                color: #c9d1d9;
            }
            QLabel {
                color: #c9d1d9;
            }
            QGroupBox {
                background-color: #21262d;
                border: 1px solid #30363d;
                border-radius: 6px;
                margin-top: 12px;
                padding: 15px;
                font-family: Consolas;
                font-weight: bold;
                color: #f79000;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #f79000;
            }
            QTableWidget {
                background-color: #161b22;
                border: 1px solid #30363d;
                gridline-color: #30363d;
                color: #c9d1d9;
                font-family: Consolas;
                font-size: 10pt;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QTableWidget::item:selected {
                background-color: #388bfd;
            }
            QHeaderView::section {
                background-color: #21262d;
                color: #f79000;
                padding: 8px;
                border: 1px solid #30363d;
                font-family: Consolas;
                font-weight: bold;
            }
            QPushButton {
                background-color: #238636;
                color: white;
                border: none;
                padding: 12px 24px;
                font-family: Consolas;
                font-size: 11pt;
                font-weight: bold;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #2ea043;
            }
            QPushButton:pressed {
                background-color: #1a7f37;
            }
            QTextEdit {
                background-color: #21262d;
                border: 1px solid #30363d;
                border-radius: 4px;
                color: #c9d1d9;
                font-family: Consolas;
                font-size: 10pt;
            }
            QScrollBar:vertical {
                background-color: #161b22;
                width: 12px;
            }
            QScrollBar::handle:vertical {
                background-color: #30363d;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #484f58;
            }
        """)
        
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(25, 25, 25, 25)
        layout.setSpacing(15)
        
        # Header
        header_label = QtWidgets.QLabel(">>> SAP ACCOUNT CREATION PREVIEW")
        header_label.setStyleSheet("font-family: Consolas; font-size: 16pt; font-weight: bold; color: #f79000;")
        layout.addWidget(header_label)
        
        subtitle_label = QtWidgets.QLabel("// Review and confirm account creation batch")
        subtitle_label.setStyleSheet("font-family: Consolas; font-size: 9pt; color: #7d8590;")
        layout.addWidget(subtitle_label)
        
        # Empty Other Description Warning
        if self.empty_other_desc:
            warning_group = QtWidgets.QGroupBox("⚠️ EMPTY_OTHER_DESCRIPTION_WARNING")
            warning_layout = QtWidgets.QVBoxLayout(warning_group)
            warning_text = QtWidgets.QTextEdit()
            warning_text.setReadOnly(True)
            warning_text.setMaximumHeight(100)
            warning_text.setStyleSheet("background-color: #3d2f1f; color: #ffa657;")
            warning_text.setText(
                f"⚠️ EMPTY OTHER DESCRIPTION: {', '.join(self.empty_other_desc)}\n\n"
                f"These {len(self.empty_other_desc)} employee ID(s) have no value in '其他說明（Other Description）'.\n"
                f"The process will continue, but please verify if this is intentional."
            )
            warning_layout.addWidget(warning_text)
            layout.addWidget(warning_group)
        
        # Existing Accounts Skipped
        existing_group = QtWidgets.QGroupBox("⚠️ EXISTING_ACCOUNTS_SKIPPED")
        existing_layout = QtWidgets.QVBoxLayout(existing_group)
        existing_text = QtWidgets.QTextEdit()
        existing_text.setReadOnly(True)
        existing_text.setMaximumHeight(100)
        
        if self.already_created:
            existing_text.setStyleSheet("background-color: #3d1f1f; color: #ff6b6b;")
            existing_text.setText(
                f"❌ DUPLICATES FOUND: {', '.join(self.already_created)}\n\n"
                f"These {len(self.already_created)} employee ID(s) already exist in the consolidated Excel.\n"
                f"They will NOT be inserted again."
            )
        else:
            existing_text.setStyleSheet("background-color: #21262d; color: #7d8590;")
            existing_text.setText("✅ [NONE] - All employee IDs are new.")
        
        existing_layout.addWidget(existing_text)
        layout.addWidget(existing_group)
        
        # New Accounts Table
        table_group = QtWidgets.QGroupBox("NEW_ACCOUNTS_QUEUE")
        table_layout = QtWidgets.QVBoxLayout(table_group)
        
        # Preview columns (exclude STATUS for new accounts)
        preview_columns = [col for col in SAP_COLUMNS if col != "STATUS"]
        
        self.table = QtWidgets.QTableWidget()
        self.table.setColumnCount(len(preview_columns))
        self.table.setHorizontalHeaderLabels(preview_columns)
        self.table.setRowCount(len(self.rows_to_append))
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.horizontalHeader().setStretchLastSection(True)
        
        # Set column widths
        for col_idx, col in enumerate(preview_columns):
            if col in ("其他說明（Other Description）", "CM remark"):
                self.table.setColumnWidth(col_idx, 260)
            elif col in ("帳號名稱（Account Name）", "郵箱（E-mail）"):
                self.table.setColumnWidth(col_idx, 220)
            else:
                self.table.setColumnWidth(col_idx, 140)
        
        # Populate table
        for row_idx, row_data in enumerate(self.rows_to_append):
            for col_idx, col in enumerate(preview_columns):
                value = row_data.get(col, "")
                item = QtWidgets.QTableWidgetItem(str(value))
                item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(row_idx, col_idx, item)
        
        table_layout.addWidget(self.table)
        layout.addWidget(table_group, 1)  # Stretch
        
        # Summary
        if len(self.rows_to_append) > 0:
            summary_text = f"✅ {len(self.rows_to_append)} NEW account(s) queued for creation"
            summary_color = "#58a6ff"
        else:
            summary_text = "⚠️ NO NEW accounts to create (all duplicates)"
            summary_color = "#ff6b6b"
        
        summary_label = QtWidgets.QLabel(summary_text)
        summary_label.setStyleSheet(f"font-family: Consolas; font-size: 10pt; font-weight: bold; color: {summary_color};")
        layout.addWidget(summary_label)
        
        # Status label for processing
        self.status_label = QtWidgets.QLabel("")
        self.status_label.setStyleSheet("font-family: Consolas; font-size: 10pt; font-weight: bold; color: #f79000;")
        layout.addWidget(self.status_label)
        
        # Execute button
        self.execute_btn = QtWidgets.QPushButton("[EXECUTE] PROCESS_BATCH")
        self.execute_btn.clicked.connect(self._on_execute)
        layout.addWidget(self.execute_btn)
        
        # Center on screen
        self.resize(1300, 900)
        screen = QtWidgets.QApplication.primaryScreen().geometry()
        self.move(
            (screen.width() - self.width()) // 2,
            (screen.height() - self.height()) // 2
        )
    
    def _on_execute(self):
        """Handle execute button click."""
        if not self.rows_to_append:
            QtWidgets.QMessageBox.warning(
                self,
                "QUEUE_EMPTY",
                "No accounts selected for creation."
            )
            return
        
        # Warn about empty descriptions
        if self.empty_other_desc:
            result = QtWidgets.QMessageBox.warning(
                self,
                "⚠️ EMPTY_OTHER_DESCRIPTION_WARNING",
                f"⚠️ WARNING: {len(self.empty_other_desc)} employee ID(s) have no value in '其他說明（Other Description）':\n\n"
                f"{', '.join(self.empty_other_desc)}\n\n"
                f"Do you still want to proceed?",
                QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
                QtWidgets.QMessageBox.StandardButton.No
            )
            if result != QtWidgets.QMessageBox.StandardButton.Yes:
                return
        
        # Update status
        self.status_label.setText("[PROCESSING] Updating consolidated database...")
        self.execute_btn.setEnabled(False)
        QtWidgets.QApplication.processEvents()
        
        # Try to save to consolidated Excel
        try:
            append_to_consolidated(self.cons_path, self.rows_to_append)
            self.status_label.setText("")
            QtWidgets.QMessageBox.information(
                self,
                "BATCH_SUCCESS",
                f"[SUCCESS] {len(self.rows_to_append)} account(s) processed successfully."
            )
        except PermissionError:
            self.status_label.setText("")
            self.execute_btn.setEnabled(True)
            QtWidgets.QMessageBox.critical(
                self,
                "FILE_LOCKED",
                f"Cannot save to consolidated Excel file.\n\n"
                f"Please close the file in Excel and try again.\n\n"
                f"File: {self.cons_path}"
            )
            return
        except Exception as e:
            self.status_label.setText("")
            self.execute_btn.setEnabled(True)
            QtWidgets.QMessageBox.critical(
                self,
                "SAVE_ERROR",
                f"Error saving to consolidated Excel:\n\n{str(e)}"
            )
            return
        
        # Prompt for ticket details
        try:
            ticket_no, ticket_img_path = prompt_ticket_details(self)
        except ValueError as exc:
            QtWidgets.QMessageBox.critical(self, "INPUT_ERROR", str(exc))
            self.execute_btn.setEnabled(True)
            return
        
        # Send email
        employee_nos = [row.get("工號（Employee No）", "") for row in self.rows_to_append]
        email_attach_paths = {
            "user_file": self.user_excel_path,
            "ticket_image": ticket_img_path if ticket_img_path else None,
            "employee_nos": employee_nos,
            "other_description_values": self.other_desc_map,
            "ticket_no": ticket_no,
            "follow_text": _extract_follow_text(self.other_desc_map, employee_nos),
        }
        send_sap_creation_email(email_attach_paths)
        
        self.accept()


def build_preview_window(
    rows_to_append: List[Dict[str, str]],
    already_created: List[str],
    cons_path: str,
    user_excel_path: str,
    other_desc_map: Dict[str, str],
) -> None:
    """Build and show the SAP preview window using PySide6."""
    # Ensure QApplication exists
    app = QtWidgets.QApplication.instance()
    if app is None:
        app = QtWidgets.QApplication([])
    
    dialog = SapPreviewDialog(
        rows_to_append=rows_to_append,
        already_created=already_created,
        cons_path=cons_path,
        user_excel_path=user_excel_path,
        other_desc_map=other_desc_map,
    )
    dialog.exec()


@dataclass
class DisableResult:
    updated: List[str]
    not_found: List[str]


def disable_sap_accounts(cons_path: str, employee_numbers: List[str]) -> DisableResult:
    """
    Disable SAP accounts by setting STATUS column to 'Disabled'.
    Checks both SR V9 file and LY V10 file sheets.
    
    Returns:
        DisableResult with lists of updated and not found employee numbers
    """
    updated: List[str] = []
    not_found: List[str] = []
    
    # Normalize input employee numbers
    emp_lookup = {str(emp).strip().upper(): str(emp).strip() for emp in employee_numbers}
    remaining = set(emp_lookup.keys())
    
    # Load workbook
    wb = load_workbook(cons_path)
    
    # Check both possible sheet names
    for sheet_name in ["SR V9 file", "LY V10 file"]:
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        headers = _get_sheet_headers(ws)
        canonical_headers = _get_sheet_canonical_headers(ws)
        
        # Find Employee No column index
        emp_col_idx: Optional[int] = None
        for idx, canonical in enumerate(canonical_headers):
            if _normalize_column_name(canonical) == _normalize_column_name("工號（Employee No）"):
                emp_col_idx = idx + 1  # openpyxl is 1-indexed
                break
        
        if emp_col_idx is None:
            continue
        
        # Find or add STATUS column
        status_col_idx: Optional[int] = None
        for idx, canonical in enumerate(canonical_headers):
            if _normalize_column_name(canonical) == _normalize_column_name("STATUS"):
                status_col_idx = idx + 1  # openpyxl is 1-indexed
                break
        
        # If STATUS column doesn't exist, add it
        if status_col_idx is None:
            status_col_idx = len(canonical_headers) + 1
            ws.cell(row=1, column=status_col_idx, value="STATUS")
            ws.cell(row=1, column=status_col_idx).alignment = Alignment(horizontal='center', vertical='center')
        
        # Iterate through rows to find employee numbers
        for row_idx in range(2, ws.max_row + 1):
            emp_value = ws.cell(row=row_idx, column=emp_col_idx).value
            if emp_value is None:
                continue
            
            emp_str = str(emp_value).strip().upper()
            if emp_str in remaining:
                # Update STATUS to Disabled
                ws.cell(row=row_idx, column=status_col_idx, value="Disabled")
                ws.cell(row=row_idx, column=status_col_idx).alignment = Alignment(horizontal='center', vertical='center')
                
                updated.append(emp_lookup[emp_str])
                remaining.remove(emp_str)
    
    # Save workbook
    wb.save(cons_path)
    wb.close()
    
    # Remaining employee numbers were not found
    not_found = [emp_lookup[emp] for emp in remaining]
    
    return DisableResult(updated=updated, not_found=not_found)
