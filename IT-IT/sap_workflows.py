"""SAP workflow helpers: data extraction, preview dialog, and email triggers."""

from __future__ import annotations

import os
import re
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk

from config_manager import get_path, set_path
from email_service import send_sap_creation_email
from openpyxl import load_workbook
from openpyxl.styles import Alignment


SAP_COLUMNS: List[str] = [
    "帳號類型（Account Type）",
    "帳號名稱（Account Name）",
    "費用代碼（Expense Code）",
    "Name",
    "聯繫電話（Contact Phone）",
    "部門（Department）",
    "工號（Employee No）",
    "郵箱（E-mail）",
    "帳號Role（Account Role）",
    "其他說明（Other Description）",
    "CM remark",
    "SR V9 file",
    "STATUS",
]

SAP_COLUMN_ALIASES: Dict[str, List[str]] = {
    "帳號類型（Account Type）": [
        "帳號類型(accounttype)",
        "accounttype",
        "type",
    ],
    "帳號名稱（Account Name）": [
        "帳號名稱(accountname)",
        "帳號名稱(account name)",
        "accountname",
        "帳號名稱",
    ],
    "費用代碼（Expense Code）": [
        "費用代碼(expensecode)",
        "費用代碼(expense code)",
        "expensecode",
        "費用代碼",
    ],
    "Name": [
        "name",
        "username",
        "fullname",
    ],
    "聯繫電話（Contact Phone）": [
        "聯繫電話(contactphone)",
        "聯繫電話(contact phone)",
        "contactphone",
        "電話",
        "phone",
    ],
    "部門（Department）": [
        "部門(department)",
        "department",
        "部門",
    ],
    "工號（Employee No）": [
        "工號(employeeno)",
        "工號(employee no)",
        "employeeno",
        "employeeid",
        "employee no",
        "employeeid",
        "empid",
    ],
    "郵箱（E-mail）": [
        "郵箱(e-mail)",
        "郵箱(email)",
        "email",
        "郵箱",
    ],
    "帳號Role（Account Role）": [
        "帳號role(accountrole)",
        "帳號role(account role)",
        "accountrole",
        "role",
    ],
    "其他說明（Other Description）": [
        "其他說明(otherdescription)",
        "其他說明(other description)",
        "otherdescription",
        "remarks",
        "備註",
    ],
    "CM remark": [
        "cmremark",
        "cm備註",
        "cm remark",
    ],
    "SR V9 file": [
        "srv9file",
        "sr v9 file",
        "srv9",
    ],
}


def _normalize_column_name(name: Optional[str]) -> str:
    if not name:
        return ""
    normalized = str(name)
    normalized = normalized.replace(" ", "").replace("\u3000", "")
    normalized = normalized.replace("（", "(").replace("）", ")")
    normalized = normalized.replace("【", "(").replace("】", ")")
    normalized = normalized.lower()
    return normalized


def _canonical_column(name: Optional[str]) -> Optional[str]:
    normalized = _normalize_column_name(name)
    if not normalized:
        return None
    for canonical, aliases in SAP_COLUMN_ALIASES.items():
        canonical_norm = _normalize_column_name(canonical)
        if normalized == canonical_norm or normalized in aliases:
            return canonical
    # If not matched, return original stripped string
    return str(name).strip() if name else None


def _build_normalized_row(row: pd.Series) -> Dict[str, any]:
    norm_map: Dict[str, any] = {}
    for col in row.index:
        norm_key = _normalize_column_name(col)
        if not norm_key:
            continue
        norm_map[norm_key] = row[col]
    return norm_map


def _get_value_from_row(norm_row: Dict[str, any], canonical_key: str) -> str:
    aliases = SAP_COLUMN_ALIASES.get(canonical_key, [])
    candidates = [_normalize_column_name(canonical_key)] + aliases
    for alias in candidates:
        if alias in norm_row:
            value = norm_row[alias]
            if pd.isna(value):
                continue
            return str(value).strip()
    return ""


def _get_sheet_headers(ws) -> List[str]:
    headers: List[str] = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value).strip() if cell.value is not None else "")
    return headers


def _get_sheet_canonical_headers(ws) -> List[str]:
    headers = _get_sheet_headers(ws)
    canonical_headers: List[str] = []
    for header in headers:
        canonical = _canonical_column(header) or header
        canonical_headers.append(canonical)
    return canonical_headers


def _get_employee_ids_from_sheet(ws) -> List[str]:
    """Extract all employee IDs from a worksheet (works with read-only mode)."""
    headers = _get_sheet_headers(ws)
    canonical_headers = _get_sheet_canonical_headers(ws)
    emp_index: Optional[int] = None
    
    # Find the Employee No column index
    for idx, canonical in enumerate(canonical_headers):
        if _normalize_column_name(canonical) == _normalize_column_name("工號（Employee No）"):
            emp_index = idx
            break
    
    if emp_index is None:
        return []
    
    # Use iter_rows to iterate through rows (compatible with read-only mode)
    ids: List[str] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        # Skip header row (min_row=2), get value from employee column
        if row_idx >= 0 and emp_index < len(row):
            value = str(row[emp_index]).strip() if row[emp_index] is not None else ""
            if value and value.lower() != "none":
                ids.append(value.upper())
    
    return ids


def _build_row_values(canonical_headers: List[str], row_data: Dict[str, str]) -> List[str]:
    values: List[str] = []
    for header in canonical_headers:
        canonical = _canonical_column(header) or header
        key = canonical
        if canonical in SAP_COLUMNS:
            key = canonical
        values.append(row_data.get(key, ""))
    return values


@dataclass
class ParsedSapData:
    rows_to_append: List[Dict[str, str]]
    already_created: List[str]
    other_desc_map: Dict[str, str]


def load_consolidated_dataframe(cons_path: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Load both sheets from consolidated Excel file.
    
    Returns:
        Tuple of (sr_v9_df, ly_v10_df)
    """
    if os.path.exists(cons_path):
        try:
            # Try to load both sheets
            sr_v9_df = pd.read_excel(cons_path, sheet_name="SR V9 file", engine="openpyxl")
        except:
            sr_v9_df = pd.DataFrame(columns=SAP_COLUMNS)
        
        try:
            ly_v10_df = pd.read_excel(cons_path, sheet_name="LY V10 file", engine="openpyxl")
        except:
            ly_v10_df = pd.DataFrame(columns=SAP_COLUMNS)
        
        return sr_v9_df, ly_v10_df
    
    return pd.DataFrame(columns=SAP_COLUMNS), pd.DataFrame(columns=SAP_COLUMNS)


def get_all_existing_employees(cons_path: str) -> List[str]:
    """
    Get all existing employee IDs from all sheets using openpyxl directly.
    Checks SR V9 file and/or LY V10 file sheets (handles if only one exists).
    
    Returns:
        List of all employee IDs found in the consolidated Excel
    """
    if not os.path.exists(cons_path):
        return []
    
    wb = load_workbook(cons_path, read_only=True)
    existing_emp = []
    
    # Check both possible sheet names (may have one or both)
    for sheet_name in ["SR V9 file", "LY V10 file"]:
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        emp_ids = _get_employee_ids_from_sheet(ws)
        existing_emp.extend(emp_ids)
    
    wb.close()
    return existing_emp


def append_to_consolidated(cons_path: str, rows: List[Dict[str, str]]) -> None:
    """
    Append new rows to the appropriate sheet.
    Preserves original Excel formatting, colors, and column widths.
    Applies center alignment to new cells.
    
    Logic:
    - If "SR V9 file" sheet exists: append rows with SR V9 value there
    - If "LY V10 file" sheet exists: append rows without SR V9 value there
    - If only one sheet exists: append all rows to that sheet
    """
    # Load existing workbook to preserve formatting
    wb = load_workbook(cons_path)
    
    # Helper to append rows with alignment
    def append_rows_to_sheet(ws, rows_data):
        # Get sheet headers to determine column order
        canonical_headers = _get_sheet_canonical_headers(ws)
        
        for row_data in rows_data:
            # Build row values in the exact order of sheet columns
            values = _build_row_values(canonical_headers, row_data)
            
            # Append the row
            start_row = ws.max_row + 1
            ws.append(values)
            
            # Apply center alignment to all cells in the new row
            for col_idx in range(1, len(values) + 1):
                cell = ws.cell(row=start_row, column=col_idx)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Check which sheets exist
    has_sr_v9 = "SR V9 file" in wb.sheetnames
    has_ly_v10 = "LY V10 file" in wb.sheetnames
    
    if has_sr_v9 and has_ly_v10:
        # Both sheets exist - separate rows based on SR V9 file column
        sr_v9_rows = []
        ly_v10_rows = []
        
        for row in rows:
            sr_v9_value = str(row.get("SR V9 file", "")).strip()
            if sr_v9_value and sr_v9_value.lower() != "nan":
                sr_v9_rows.append(row)
            else:
                ly_v10_rows.append(row)
        
        if sr_v9_rows:
            append_rows_to_sheet(wb["SR V9 file"], sr_v9_rows)
        if ly_v10_rows:
            append_rows_to_sheet(wb["LY V10 file"], ly_v10_rows)
    
    elif has_ly_v10:
        # Only LY V10 sheet exists - append all rows there
        append_rows_to_sheet(wb["LY V10 file"], rows)
    
    elif has_sr_v9:
        # Only SR V9 sheet exists - append all rows there
        append_rows_to_sheet(wb["SR V9 file"], rows)
    
    else:
        raise ValueError("No valid sheet found in consolidated Excel (expected 'SR V9 file' or 'LY V10 file')")
    
    # Save workbook - this preserves all formatting, colors, column widths
    wb.save(cons_path)


def parse_user_excel(user_df: pd.DataFrame, existing_emp: List[str]) -> ParsedSapData:
    existing_set = set(str(emp).strip().upper() for emp in existing_emp if str(emp).strip())

    rows_to_append: List[Dict[str, str]] = []
    already_created: List[str] = []
    other_desc_map: Dict[str, str] = {}

    for _, row in user_df.iterrows():
        # Build normalized row for flexible column matching
        norm_row = _build_normalized_row(row)
        
        # Get employee number using normalized matching
        emp_no = _get_value_from_row(norm_row, "工號（Employee No）")
        if not emp_no:
            continue
        
        # Check if employee already exists (case-insensitive)
        if emp_no.upper() in existing_set:
            already_created.append(emp_no)
            continue

        # Build new row using normalized column matching for all fields
        new_row = {
            "帳號類型（Account Type）": _get_value_from_row(norm_row, "帳號類型（Account Type）"),
            "帳號名稱（Account Name）": _get_value_from_row(norm_row, "帳號名稱（Account Name）"),
            "費用代碼（Expense Code）": _get_value_from_row(norm_row, "費用代碼（Expense Code）"),
            "Name": _get_value_from_row(norm_row, "Name"),
            "聯繫電話（Contact Phone）": _get_value_from_row(norm_row, "聯繫電話（Contact Phone）"),
            "部門（Department）": _get_value_from_row(norm_row, "部門（Department）"),
            "工號（Employee No）": emp_no,
            "郵箱（E-mail）": _get_value_from_row(norm_row, "郵箱（E-mail）"),
            "帳號Role（Account Role）": _get_value_from_row(norm_row, "帳號Role（Account Role）"),
            "其他說明（Other Description）": _get_value_from_row(norm_row, "其他說明（Other Description）"),
            "CM remark": _get_value_from_row(norm_row, "CM remark"),
            "SR V9 file": _get_value_from_row(norm_row, "SR V9 file"),
        }
        rows_to_append.append(new_row)
        other_desc_map[emp_no] = str(new_row["其他說明（Other Description）"]).strip()

    return ParsedSapData(rows_to_append, already_created, other_desc_map)


def _extract_follow_text(other_desc_map: Dict[str, str], employee_nos: List[str]) -> str:
    if not employee_nos:
        return ""
    pattern = re.compile(r"(SGP\d+)", re.IGNORECASE)
    for emp in employee_nos:
        match = pattern.search(other_desc_map.get(emp, ""))
        if match:
            return f"權限跟隨{match.group(1)}即可"
    return ""


def prompt_ticket_details(parent: tk.Toplevel) -> Tuple[str, str]:
    dialog_kwargs = {
        "title": "Select Ticket Image File",
        "filetypes": [("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif"), ("All files", "*.*")],
    }
    default_ticket_dir = get_path("sap_ticket_image_dir")
    if default_ticket_dir and os.path.isdir(default_ticket_dir):
        dialog_kwargs["initialdir"] = default_ticket_dir

    ticket_img_path = filedialog.askopenfilename(parent=parent, **dialog_kwargs)
    if ticket_img_path:
        set_path("sap_ticket_image_dir", os.path.dirname(ticket_img_path))

    ticket_no = simpledialog.askstring("Ticket Number", "Enter ticket number (e.g. S0000YE9G):", parent=parent)
    if not ticket_no:
        raise ValueError("Ticket number is required")

    return ticket_no, ticket_img_path


def build_preview_window(
    rows_to_append: List[Dict[str, str]],
    already_created: List[str],
    cons_path: str,
    user_excel_path: str,
    other_desc_map: Dict[str, str],
) -> None:
    preview = tk.Toplevel()
    preview.title("SAP_PREVIEW_MODULE")
    
    # Calculate responsive size based on screen dimensions
    screen_width = preview.winfo_screenwidth()
    screen_height = preview.winfo_screenheight()
    
    # Use 70% of screen size, but cap at reasonable maximums
    window_width = min(int(screen_width * 0.7), 1400)
    window_height = min(int(screen_height * 0.7), 900)
    
    # Center the window
    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2
    
    preview.geometry(f"{window_width}x{window_height}+{x}+{y}")
    preview.configure(bg="#0d1117")

    style = ttk.Style(preview)
    style.configure("Preview.TFrame", background="#0d1117", relief="flat")
    style.configure(
        "Preview.TLabelframe",
        background="#21262d",
        foreground="#f79000",
        borderwidth=1,
        relief="solid",
        padding=15,
    )
    style.configure(
        "Preview.TLabelframe.Label",
        background="#21262d",
        foreground="#f79000",
        font=("Consolas", 11, "bold"),
    )

    # Create scrollable canvas
    canvas = tk.Canvas(preview, bg="#0d1117", highlightthickness=0)
    scrollbar = ttk.Scrollbar(preview, orient="vertical", command=canvas.yview)
    scrollable_frame = ttk.Frame(canvas, style="Preview.TFrame")
    
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    
    # Mouse wheel scrolling for canvas (but not for treeview)
    def _on_canvas_mousewheel(event):
        # Only scroll canvas if mouse is not over the treeview
        widget = event.widget
        # Check if the event is from the treeview or its children
        if hasattr(widget, 'winfo_class') and widget.winfo_class() == 'Treeview':
            return  # Don't scroll canvas, let treeview handle it
        canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
    
    # Bind to canvas and scrollable_frame, not all widgets
    canvas.bind("<MouseWheel>", _on_canvas_mousewheel)
    scrollable_frame.bind("<MouseWheel>", _on_canvas_mousewheel)

    main_frame = ttk.Frame(scrollable_frame, style="Preview.TFrame", padding="25")
    main_frame.pack(fill="both", expand=True)

    header_frame = ttk.Frame(main_frame, style="Preview.TFrame")
    header_frame.pack(fill="x", pady=(0, 25))
    ttk.Label(
        header_frame,
        text=">>> SAP ACCOUNT CREATION PREVIEW",
        font=("Consolas", 16, "bold"),
        background="#0d1117",
        foreground="#f79000",
        anchor="w",
    ).pack(fill="x")
    ttk.Label(
        header_frame,
        text="// Review and confirm account creation batch",
        font=("Consolas", 9),
        background="#0d1117",
        foreground="#7d8590",
        anchor="w",
    ).pack(fill="x", pady=(5, 0))

    # Check for empty "其他說明（Other Description）" fields
    empty_other_desc = []
    for row in rows_to_append:
        emp_no = row.get("工號（Employee No）", "")
        other_desc = str(row.get("其他說明（Other Description）", "")).strip()
        if not other_desc or other_desc.lower() == "nan":
            empty_other_desc.append(emp_no)
    
    # Show warning for empty Other Description
    if empty_other_desc:
        other_desc_frame = ttk.LabelFrame(main_frame, text="⚠️ EMPTY_OTHER_DESCRIPTION_WARNING", style="Preview.TLabelframe")
        other_desc_frame.pack(fill="x", pady=(0, 15))
        other_desc_content = ttk.Frame(other_desc_frame, style="Preview.TFrame")
        other_desc_content.pack(fill="x", padx=10, pady=10)
        
        bg_color = "#3d2f1f"  # Dark orange background
        fg_color = "#ffa657"  # Light orange text
        display_text = f"⚠️ EMPTY OTHER DESCRIPTION: {', '.join(empty_other_desc)}\n\n" \
                      f"These {len(empty_other_desc)} employee ID(s) have no value in '其他說明（Other Description）'.\n" \
                      f"The process will continue, but please verify if this is intentional."
        
        other_desc_text = tk.Text(
            other_desc_content,
            font=("Consolas", 10),
            bg=bg_color,
            fg=fg_color,
            height=4,
            wrap="word",
            relief="flat",
        )
        other_desc_text.insert("1.0", display_text)
        other_desc_text.configure(state="disabled")
        other_desc_text.pack(fill="both", expand=True)
    
    # Show skipped accounts prominently
    existing_frame = ttk.LabelFrame(main_frame, text="⚠️ EXISTING_ACCOUNTS_SKIPPED", style="Preview.TLabelframe")
    existing_frame.pack(fill="x", pady=(0, 15))
    existing_content = ttk.Frame(existing_frame, style="Preview.TFrame")
    existing_content.pack(fill="x", padx=10, pady=10)
    
    if already_created:
        # Show in red/warning color if there are duplicates
        bg_color = "#3d1f1f"  # Dark red background
        fg_color = "#ff6b6b"  # Light red text
        display_text = f"❌ DUPLICATES FOUND: {', '.join(already_created)}\n\n" \
                      f"These {len(already_created)} employee ID(s) already exist in the consolidated Excel.\n" \
                      f"They will NOT be inserted again."
    else:
        bg_color = "#21262d"
        fg_color = "#7d8590"
        display_text = "✅ [NONE] - All employee IDs are new."
    
    existing_text = tk.Text(
        existing_content,
        font=("Consolas", 10),
        bg=bg_color,
        fg=fg_color,
        height=4,
        wrap="word",
        relief="flat",
    )
    existing_text.insert("1.0", display_text)
    existing_text.configure(state="disabled")
    existing_text.pack(fill="both", expand=True)

    new_frame = ttk.LabelFrame(main_frame, text="NEW_ACCOUNTS_QUEUE", style="Preview.TLabelframe")
    new_frame.pack(fill="both", expand=True, pady=(0, 15))
    tree_container = ttk.Frame(new_frame, style="Preview.TFrame")
    tree_container.pack(fill="both", expand=True, padx=15, pady=10)

    # Create vertical scrollbar
    y_scroll = ttk.Scrollbar(tree_container, orient="vertical")
    
    # Create horizontal scrollbar
    x_scroll = ttk.Scrollbar(tree_container, orient="horizontal")

    # Show ALL columns in preview in exact order as consolidated Excel
    preview_columns = SAP_COLUMNS
    tree = ttk.Treeview(
        tree_container,
        columns=preview_columns,
        show="headings",
        height=12,
        selectmode="extended",
        yscrollcommand=y_scroll.set,
        xscrollcommand=x_scroll.set,
    )

    # Configure headings and column alignment/width
    for col in preview_columns:
        tree.heading(col, text=col)
        anchor = "center"
        width = 160
        if col in ("其他說明（Other Description）", "CM remark"):
            anchor = "w"
            width = 260
        if col in ("帳號名稱（Account Name）", "郵箱（E-mail）"):
            width = 220
        tree.column(col, anchor=anchor, width=width, stretch=False, minwidth=100)

    # Pack with proper scroll configuration
    tree.grid(row=0, column=0, sticky="nsew")
    y_scroll.grid(row=0, column=1, sticky="ns")
    x_scroll.grid(row=1, column=0, sticky="ew")
    
    # Configure grid weights for proper resizing
    tree_container.grid_rowconfigure(0, weight=1)
    tree_container.grid_columnconfigure(0, weight=1)
    
    # Link scrollbars to tree
    y_scroll.configure(command=tree.yview)
    x_scroll.configure(command=tree.xview)
    
    # Enable mouse wheel scrolling (stop propagation to prevent canvas scroll)
    def _on_tree_mousewheel(event):
        tree.yview_scroll(int(-1 * (event.delta / 120)), "units")
        return "break"  # Stop event propagation
    
    def _on_tree_shift_mousewheel(event):
        tree.xview_scroll(int(-1 * (event.delta / 120)), "units")
        return "break"  # Stop event propagation
    
    tree.bind("<MouseWheel>", _on_tree_mousewheel)
    tree.bind("<Shift-MouseWheel>", _on_tree_shift_mousewheel)
    tree_container.bind("<MouseWheel>", _on_tree_mousewheel)
    tree_container.bind("<Shift-MouseWheel>", _on_tree_shift_mousewheel)

    # Populate rows in the same mapped order
    checks: List[Tuple[str, Dict[str, str]]] = []
    for row in rows_to_append:
        values = [
            row.get("帳號類型（Account Type）", ""),
            row.get("帳號名稱（Account Name）", ""),
            row.get("費用代碼（Expense Code）", ""),
            row.get("Name", ""),
            row.get("聯繫電話（Contact Phone）", ""),
            row.get("部門（Department）", ""),
            row.get("工號（Employee No）", ""),
            row.get("郵箱（E-mail）", ""),
            row.get("帳號Role（Account Role）", ""),
            row.get("其他說明（Other Description）", ""),
            row.get("CM remark", ""),
            row.get("SR V9 file", ""),
        ]
        iid = tree.insert("", "end", values=values)
        checks.append((iid, row))

    # Summary label
    if len(checks) > 0:
        summary_text = f"✅ {len(checks)} NEW account(s) queued for creation"
        summary_color = "#58a6ff"
    else:
        summary_text = f"⚠️ NO NEW accounts to create (all duplicates)"
        summary_color = "#ff6b6b"
    
    ttk.Label(
        main_frame,
        text=summary_text,
        font=("Consolas", 10, "bold"),
        background="#0d1117",
        foreground=summary_color,
        anchor="w",
    ).pack(fill="x", pady=(0, 15))

    def confirm() -> None:
        selected_rows = [row for _, row in checks]
        employee_nos = [row.get("工號（Employee No）", "") for _, row in checks]
        if not selected_rows:
            messagebox.showwarning("QUEUE_EMPTY", "No accounts selected for creation.", parent=preview)
            return
        
        # Check if any rows have empty "其他說明（Other Description）"
        if empty_other_desc:
            proceed = messagebox.askyesno(
                "⚠️ EMPTY_OTHER_DESCRIPTION_WARNING",
                f"⚠️ WARNING: {len(empty_other_desc)} employee ID(s) have no value in '其他說明（Other Description）':\n\n"
                f"{', '.join(empty_other_desc)}\n\n"
                f"Do you still want to proceed?",
                parent=preview,
                icon='warning'
            )
            if not proceed:
                return

        indicator = ttk.Label(
            main_frame,
            text="[PROCESSING] Updating consolidated database...",
            font=("Consolas", 10, "bold"),
            background="#0d1117",
            foreground="#f79000",
        )
        indicator.pack(pady=5)
        preview.update()

        try:
            append_to_consolidated(cons_path, selected_rows)
            indicator.destroy()
            messagebox.showinfo(
                "BATCH_SUCCESS",
                f"[SUCCESS] {len(selected_rows)} account(s) processed successfully.",
                parent=preview,
            )
        except PermissionError:
            indicator.destroy()
            messagebox.showerror(
                "FILE_LOCKED",
                f"Cannot save to consolidated Excel file.\n\n"
                f"Please close the file in Excel and try again.\n\n"
                f"File: {cons_path}",
                parent=preview
            )
            return
        except Exception as e:
            indicator.destroy()
            messagebox.showerror(
                "SAVE_ERROR",
                f"Error saving to consolidated Excel:\n\n{str(e)}",
                parent=preview
            )
            return

        try:
            ticket_no, ticket_img_path = prompt_ticket_details(preview)
        except ValueError as exc:
            messagebox.showerror("INPUT_ERROR", str(exc), parent=preview)
            return

        email_attach_paths = {
            "user_file": user_excel_path,
            "ticket_image": ticket_img_path if ticket_img_path else None,
            "employee_nos": employee_nos,
            "other_description_values": other_desc_map,
            "ticket_no": ticket_no,
            "follow_text": _extract_follow_text(other_desc_map, employee_nos),
        }
        send_sap_creation_email(email_attach_paths)
        preview.destroy()

    ttk.Button(main_frame, text="[EXECUTE] PROCESS_BATCH", command=confirm).pack(fill="x", pady=10)

    # Window is already centered and sized, just finalize
    preview.update_idletasks()
    preview.transient()
    preview.grab_set()


@dataclass
class DisableResult:
    updated: List[str]
    not_found: List[str]


def disable_sap_accounts(cons_path: str, employee_numbers: List[str]) -> DisableResult:
    """
    Disable SAP accounts by setting STATUS column to 'Disabled'.
    Checks both SR V9 file and LY V10 file sheets.
    
    Returns:
        DisableResult with lists of updated and not found employee numbers
    """
    updated: List[str] = []
    not_found: List[str] = []
    
    # Normalize input employee numbers
    emp_lookup = {str(emp).strip().upper(): str(emp).strip() for emp in employee_numbers}
    remaining = set(emp_lookup.keys())
    
    # Load workbook
    wb = load_workbook(cons_path)
    
    # Check both possible sheet names
    for sheet_name in ["SR V9 file", "LY V10 file"]:
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        headers = _get_sheet_headers(ws)
        canonical_headers = _get_sheet_canonical_headers(ws)
        
        # Find Employee No column index
        emp_col_idx: Optional[int] = None
        for idx, canonical in enumerate(canonical_headers):
            if _normalize_column_name(canonical) == _normalize_column_name("工號（Employee No）"):
                emp_col_idx = idx + 1  # openpyxl is 1-indexed
                break
        
        if emp_col_idx is None:
            continue
        
        # Find or add STATUS column
        status_col_idx: Optional[int] = None
        for idx, canonical in enumerate(canonical_headers):
            if _normalize_column_name(canonical) == _normalize_column_name("STATUS"):
                status_col_idx = idx + 1  # openpyxl is 1-indexed
                break
        
        # If STATUS column doesn't exist, add it
        if status_col_idx is None:
            status_col_idx = len(canonical_headers) + 1
            ws.cell(row=1, column=status_col_idx, value="STATUS")
            ws.cell(row=1, column=status_col_idx).alignment = Alignment(horizontal='center', vertical='center')
        
        # Iterate through rows to find employee numbers
        for row_idx in range(2, ws.max_row + 1):
            emp_value = ws.cell(row=row_idx, column=emp_col_idx).value
            if emp_value is None:
                continue
            
            emp_str = str(emp_value).strip().upper()
            if emp_str in remaining:
                # Update STATUS to Disabled
                ws.cell(row=row_idx, column=status_col_idx, value="Disabled")
                ws.cell(row=row_idx, column=status_col_idx).alignment = Alignment(horizontal='center', vertical='center')
                
                updated.append(emp_lookup[emp_str])
                remaining.remove(emp_str)
    
    # Save workbook
    wb.save(cons_path)
    wb.close()
    
    # Remaining employee numbers were not found
    not_found = [emp_lookup[emp] for emp in remaining]
    
    return DisableResult(updated=updated, not_found=not_found)


